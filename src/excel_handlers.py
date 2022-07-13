from datetime import datetime
from itertools import product
from typing import Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils.cell import get_column_letter

from data_types import ProductItem, ShopName, ProductName, InfoForDb, ProductItemForResultSheet
from config import ROOT_DIR


COMMON_CELL_STYLE = NamedStyle(
    name='COMMON_CELL_STYLE',
    alignment=Alignment(
        horizontal='center',
        vertical='center',
        indent=10
    ),
    font=Font(
        name='Arial',
        size=10
    ),
)


def delete_empty_cols(worksheet: Worksheet) -> Worksheet:
    empty_cols = []

    for col in worksheet.iter_cols():
        cell = col[0]
        if cell.value == None:
            empty_cols.append(cell.column)

    for col_index in empty_cols[::-1]:
        worksheet.delete_cols(col_index)

    return worksheet


def delete_empty_rows(worksheet: Worksheet) -> Worksheet:
    empty_rows = []

    for row in worksheet.iter_rows():
        cell = row[0]
        if cell.value == None:
            empty_rows.append(cell.row)

    for row_index in empty_rows[::-1]:
        worksheet.delete_rows(row_index)

    return worksheet


def parse_products(worksheet: Worksheet) -> Iterator[ProductName]:
    for product in worksheet.iter_rows(min_row=2, max_col=1):
        yield product[0].value


def parse_shops(worksheet: Worksheet) -> Iterator[ShopName]:
    for shop in worksheet.iter_cols(min_col=2, max_row=1):
        yield shop[0].value


def parse_product_items(
    worksheet: Worksheet,
) -> Iterator[ProductItem]:

    for column in worksheet.iter_cols(min_row=2, min_col=2):

        for cell in column:
            shop_name = worksheet.cell(column=cell.column, row=1).value
            product_name = worksheet.cell(column=1, row=cell.row).value

            yield ProductItem(
                shop_name=shop_name,
                product_name=product_name,
                url=cell.value
            )


def parce_excell_file(path_to_excell_file: str) -> InfoForDb:
    workbook = load_workbook(path_to_excell_file)
    worksheet = delete_empty_cols(delete_empty_rows(workbook.active))

    shop_names = parse_shops(worksheet)
    product_names = parse_products(worksheet)
    product_item_links = parse_product_items(worksheet)

    return InfoForDb(shop_names, product_names, product_item_links)


def create_result_sheet_header(worksheet: Worksheet, shops, products) -> Worksheet:
    BASE_COLUMN_WIDTH = 10
    TITLE_COLUMN_WIDTH = 40

    worksheet.merge_cells('A1:A2')
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 25


    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = 'Title'
    title_cell.style = COMMON_CELL_STYLE
    worksheet.column_dimensions['A'].width = TITLE_COLUMN_WIDTH

    columns_count = 1
    for shop in shops:
        worksheet.merge_cells(
            start_row=1,
            end_row=1,
            start_column=columns_count + 1,
            end_column=columns_count + 3
        )

        shop_name_cell: Cell = worksheet.cell(row=1, column=columns_count + 1)
        shop_name_cell.value = shop.name
        shop_name_cell.style = COMMON_CELL_STYLE

        url_cell = worksheet.cell(row=2, column=columns_count + 1)
        url_cell.value = "URL"
        url_cell.style = COMMON_CELL_STYLE
        worksheet.column_dimensions[get_column_letter(columns_count + 1)].width = BASE_COLUMN_WIDTH

        selling_price_cell = worksheet.cell(row=2, column=columns_count + 2)
        selling_price_cell.value = "Selling price"
        selling_price_cell.style = COMMON_CELL_STYLE
        worksheet.column_dimensions[get_column_letter(columns_count + 2)].width = BASE_COLUMN_WIDTH

        net_price_cell = worksheet.cell(row=2, column=columns_count + 3)
        net_price_cell.value = "Price before discount"
        net_price_cell.style = COMMON_CELL_STYLE
        net_price_cell.alignment = Alignment(
                                        horizontal='center',
                                        vertical='center',
                                        wrap_text=True,
                                        indent=10
                                    )

        worksheet.column_dimensions[get_column_letter(columns_count + 3)].width = BASE_COLUMN_WIDTH

        columns_count += 3

    rows_count = 3
    for product in products:
        worksheet.cell(row=rows_count, column=1).value = product.name
        worksheet.row_dimensions[rows_count].height = 20
        rows_count += 1


def fill_result_sheet(worksheet: Worksheet, product_items: list[ProductItemForResultSheet]):
    sheet_cols = tuple(worksheet.iter_cols(max_row=1, min_col=2))
    sheet_rows = tuple(worksheet.iter_rows(min_row=3, max_col=1))

    for item in product_items:
        for col, col_indx in zip(sheet_cols, range(2, len(sheet_cols) + 3)):
            if item.shop_name == col[0].value:
                for row, row_indx in zip(sheet_rows, range(3, len(sheet_rows) + 4)):
                    if item.product_name == row[0].value:
                        url_cell = worksheet.cell(column=col_indx, row=row_indx)
                        url_cell.value = item.product_item_link if item.product_item_link != None else '-'
                        url_cell.style = COMMON_CELL_STYLE

                        selling_price_cell = worksheet.cell(column=col_indx + 1, row=row_indx)
                        selling_price_cell.value = item.selling_price if item.selling_price != None else '-'
                        selling_price_cell.style = COMMON_CELL_STYLE

                        net_price_cell = worksheet.cell(column=col_indx + 2, row=row_indx)
                        net_price_cell.value = item.net_price if item.net_price != None else '-'
                        net_price_cell.style = COMMON_CELL_STYLE



if __name__ == '__main__':
    worksheet = Workbook().active
    worksheet.merge_cells('A1:B4')
    for col in worksheet.iter_cols():
        print(get_column_letter(col[0].column))


